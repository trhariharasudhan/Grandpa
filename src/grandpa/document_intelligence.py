"""Local document intelligence for Grandpa.

Supports lightweight, dependency-friendly extraction for common office files.
Everything stays local and all organization actions are dry-run by default.
"""

from __future__ import annotations

import csv
import json
import os
import re
import sqlite3
import time
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from grandpa.core.config import DEFAULT_CONFIG_DIR


DEFAULT_DOCUMENT_DB = DEFAULT_CONFIG_DIR / "document_intelligence.db"
SUPPORTED_EXTENSIONS = {".txt", ".md", ".markdown", ".log", ".csv", ".json", ".pdf", ".docx", ".xlsx", ".pptx"}
MAX_TEXT_CHARS = 20000
SKIP_DIRS = {".git", ".venv", "node_modules", "__pycache__", "build", "cache", ".cache"}


@dataclass(frozen=True)
class DocumentResult:
    status: str
    message: str
    data: dict[str, Any]


class DocumentIndex:
    def __init__(self, db_path: Path | str = DEFAULT_DOCUMENT_DB) -> None:
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_db()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self) -> None:
        with self._connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS documents (
                    path TEXT PRIMARY KEY,
                    indexed_at REAL NOT NULL,
                    title TEXT,
                    type TEXT NOT NULL,
                    modified REAL NOT NULL,
                    size INTEGER NOT NULL,
                    tags_json TEXT NOT NULL,
                    summary TEXT,
                    text_sample TEXT
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_documents_type ON documents(type)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_documents_modified ON documents(modified)")

    def upsert(self, metadata: dict[str, Any]) -> None:
        with self._connect() as conn:
            conn.execute(
                """
                INSERT INTO documents(path, indexed_at, title, type, modified, size, tags_json, summary, text_sample)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(path) DO UPDATE SET
                    indexed_at = excluded.indexed_at,
                    title = excluded.title,
                    type = excluded.type,
                    modified = excluded.modified,
                    size = excluded.size,
                    tags_json = excluded.tags_json,
                    summary = excluded.summary,
                    text_sample = excluded.text_sample
                """,
                (
                    metadata["path"],
                    time.time(),
                    metadata.get("title", ""),
                    metadata["type"],
                    metadata["modified"],
                    metadata["size"],
                    json.dumps(metadata.get("tags", [])),
                    metadata.get("summary", ""),
                    metadata.get("text_sample", ""),
                ),
            )

    def search(self, query: str, *, limit: int = 20) -> list[dict[str, Any]]:
        tokens = _tokens(query)
        rows = self.all(limit=2000)
        scored: list[tuple[int, dict[str, Any]]] = []
        for row in rows:
            haystack = " ".join(
                [
                    row.get("title", ""),
                    Path(row["path"]).name,
                    row.get("summary", ""),
                    row.get("text_sample", ""),
                    " ".join(row.get("tags", [])),
                ]
            ).lower()
            score = sum(1 for token in tokens if token in haystack)
            if _query_wants_type(query, row["type"]):
                score += 3
            if _query_wants_last_month(query, row["modified"]):
                score += 3
            if score:
                scored.append((score, row))
        scored.sort(key=lambda item: (item[0], item[1]["modified"]), reverse=True)
        return [row for _score, row in scored[:limit]]

    def all(self, *, limit: int = 100) -> list[dict[str, Any]]:
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT path, indexed_at, title, type, modified, size, tags_json, summary, text_sample
                FROM documents
                ORDER BY modified DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
        result = []
        for row in rows:
            item = {key: row[key] for key in row.keys()}
            try:
                item["tags"] = json.loads(item.pop("tags_json") or "[]")
            except json.JSONDecodeError:
                item["tags"] = []
            result.append(item)
        return result


def extract_document_text(path: Path | str) -> str:
    path = Path(path)
    suffix = path.suffix.lower()
    try:
        if suffix in {".txt", ".md", ".markdown", ".log", ".json"}:
            return path.read_text(encoding="utf-8", errors="ignore")[:MAX_TEXT_CHARS]
        if suffix == ".csv":
            return _read_csv_text(path)
        if suffix == ".pdf":
            return _read_pdf_text(path)
        if suffix == ".docx":
            return _read_docx_text(path)
        if suffix == ".xlsx":
            return _read_xlsx_text(path)
        if suffix == ".pptx":
            return _read_pptx_text(path)
    except Exception:
        return ""
    return ""


def summarize_document(path: Path | str) -> DocumentResult:
    path = Path(path)
    text = extract_document_text(path)
    if not text:
        return DocumentResult("unsupported", f"I could not extract readable text from {path.name}.", {"path": str(path)})
    metadata = document_metadata(path, text=text)
    tables = extract_tables(path, text=text)
    message = f"Summary of {path.name}:\n\n{metadata['summary']}"
    if tables:
        message += f"\n\nDetected {len(tables)} table/data block(s)."
    return DocumentResult("handled", message, {"metadata": metadata, "tables": tables})


def document_metadata(path: Path | str, *, text: str | None = None) -> dict[str, Any]:
    path = Path(path)
    stat = path.stat()
    text = text if text is not None else extract_document_text(path)
    title = _title_from_text(text) or path.stem
    tags = _tags_from_text(path.name + " " + text[:4000])
    return {
        "path": str(path),
        "name": path.name,
        "title": title[:160],
        "type": path.suffix.lower().lstrip(".") or "file",
        "modified": stat.st_mtime,
        "modified_label": datetime.fromtimestamp(stat.st_mtime).strftime("%b %d, %Y"),
        "size": stat.st_size,
        "tags": tags,
        "summary": smart_summary(text),
        "text_sample": text[:3000],
    }


def index_documents(paths: list[Path] | None = None, *, limit: int = 500) -> dict[str, Any]:
    index = DocumentIndex()
    paths = paths or list(_iter_safe_documents(limit=limit))
    indexed = 0
    skipped = 0
    for path in paths[:limit]:
        text = extract_document_text(path)
        if not text and path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            skipped += 1
            continue
        try:
            index.upsert(document_metadata(path, text=text))
            indexed += 1
        except OSError:
            skipped += 1
    return {"indexed": indexed, "skipped": skipped, "storage": str(index.db_path), "local_only": True}


def search_documents(query: str, *, limit: int = 10) -> DocumentResult:
    index_documents(limit=300)
    results = DocumentIndex().search(query, limit=limit)
    if not results:
        return DocumentResult("handled", f"I did not find local documents matching {query}.", {"results": []})
    lines = [f"Documents matching {query}:"]
    for item in results:
        lines.append(f"- {Path(item['path']).name} ({item['type']}, {item['modified_label'] if 'modified_label' in item else _date(item['modified'])})")
    return DocumentResult("handled", "\n".join(lines), {"results": results})


def suggest_renames(query: str = "", *, limit: int = 10) -> DocumentResult:
    index_documents(limit=300)
    rows = DocumentIndex().search(query, limit=limit) if query else DocumentIndex().all(limit=limit)
    suggestions = []
    for row in rows:
        path = Path(row["path"])
        stem = _slug(row.get("title") or path.stem)
        if not stem:
            continue
        suggested = f"{_date(row['modified'], compact=True)}-{stem}{path.suffix.lower()}"
        if suggested.lower() != path.name.lower():
            suggestions.append({"path": str(path), "current": path.name, "suggested": suggested, "dry_run": True})
    return DocumentResult(
        "handled",
        "Suggested safe rename dry-run:\n" + "\n".join(f"- {item['current']} -> {item['suggested']}" for item in suggestions[:limit])
        if suggestions
        else "I did not find useful rename suggestions.",
        {"suggestions": suggestions, "dry_run": True},
    )


def organization_plan(query: str = "", *, dry_run: bool = True) -> DocumentResult:
    index_documents(limit=300)
    rows = DocumentIndex().search(query, limit=25) if query else DocumentIndex().all(limit=25)
    moves = []
    for row in rows:
        path = Path(row["path"])
        category = _organization_category(row)
        target_dir = path.parent / category
        moves.append({"path": str(path), "target": str(target_dir / path.name), "category": category, "dry_run": dry_run})
    return DocumentResult(
        "requires_confirmation" if not dry_run else "handled",
        f"Prepared {'dry-run ' if dry_run else ''}organization plan with {len(moves)} item(s).",
        {"moves": moves, "dry_run": dry_run, "approval_required": not dry_run},
    )


def extract_tables(path: Path | str, *, text: str | None = None) -> list[dict[str, Any]]:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return [_csv_table(path)]
    if suffix == ".xlsx":
        return _xlsx_tables(path)
    text = text if text is not None else extract_document_text(path)
    rows = []
    for line in text.splitlines():
        if "," in line:
            cells = [cell.strip() for cell in line.split(",")]
        elif "\t" in line:
            cells = [cell.strip() for cell in line.split("\t")]
        elif re.search(r"\s{2,}", line):
            cells = [cell.strip() for cell in re.split(r"\s{2,}", line)]
        else:
            continue
        if len([cell for cell in cells if cell]) >= 2:
            rows.append(cells)
    return [{"kind": "text_table", "rows": rows[:30], "row_count": len(rows)}] if rows else []


def diagnostics() -> dict[str, Any]:
    index = DocumentIndex()
    rows = index.all(limit=1000)
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["type"]] = counts.get(row["type"], 0) + 1
    return {
        "status": "ready",
        "supported_types": sorted(ext.lstrip(".") for ext in SUPPORTED_EXTENSIONS),
        "indexed_documents": len(rows),
        "type_counts": counts,
        "storage": {"backend": "sqlite", "path": str(index.db_path), "local_only": True},
        "safety": {
            "local_only": True,
            "silent_delete": False,
            "bulk_operations_require_approval": True,
        },
    }


def _read_csv_text(path: Path) -> str:
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        rows = list(csv.reader(handle))[:80]
    return "\n".join(", ".join(cell for cell in row) for row in rows)[:MAX_TEXT_CHARS]


def _read_pdf_text(path: Path) -> str:
    try:
        import pdfplumber  # type: ignore[import-untyped]

        chunks = []
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages[:10]:
                chunks.append(page.extract_text() or "")
        return "\n".join(chunks)[:MAX_TEXT_CHARS]
    except Exception:
        return ""


def _read_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as doc:
        xml = doc.read("word/document.xml")
    root = ET.fromstring(xml)
    texts = [node.text or "" for node in root.iter() if node.tag.endswith("}t")]
    return "\n".join(texts)[:MAX_TEXT_CHARS]


def _read_pptx_text(path: Path) -> str:
    chunks: list[str] = []
    with zipfile.ZipFile(path) as deck:
        slide_names = sorted(name for name in deck.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml"))
        for name in slide_names[:30]:
            root = ET.fromstring(deck.read(name))
            texts = [node.text or "" for node in root.iter() if node.tag.endswith("}t")]
            if texts:
                chunks.append(" ".join(texts))
    return "\n".join(chunks)[:MAX_TEXT_CHARS]


def _read_xlsx_text(path: Path) -> str:
    tables = _xlsx_tables(path)
    lines: list[str] = []
    for table in tables:
        lines.append(f"Sheet: {table['sheet']}")
        lines.extend(", ".join(str(cell) for cell in row) for row in table["rows"][:30])
    return "\n".join(lines)[:MAX_TEXT_CHARS]


def _xlsx_tables(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(path, read_only=True, data_only=True)
        tables = []
        for sheet in workbook.worksheets[:8]:
            rows = []
            for row in sheet.iter_rows(max_row=60, values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(values):
                    rows.append(values)
            tables.append({"kind": "xlsx", "sheet": sheet.title, "rows": rows, "row_count": len(rows)})
        workbook.close()
        return tables
    except Exception:
        pass
    try:
        return _xlsx_tables_from_zip(path)
    except Exception:
        return []


def _xlsx_tables_from_zip(path: Path) -> list[dict[str, Any]]:
    with zipfile.ZipFile(path) as workbook:
        shared = _xlsx_shared_strings(workbook)
        sheet_names = sorted(
            name for name in workbook.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        tables = []
        for idx, name in enumerate(sheet_names[:8], start=1):
            root = ET.fromstring(workbook.read(name))
            rows = []
            for row in root.iter():
                if not row.tag.endswith("}row"):
                    continue
                values = []
                for cell in row:
                    if not cell.tag.endswith("}c"):
                        continue
                    cell_type = cell.attrib.get("t", "")
                    value = ""
                    for child in cell:
                        if child.tag.endswith("}v") and child.text is not None:
                            value = child.text
                            break
                    if cell_type == "s" and value.isdigit():
                        shared_idx = int(value)
                        value = shared[shared_idx] if shared_idx < len(shared) else value
                    values.append(value)
                if any(values):
                    rows.append(values)
            tables.append({"kind": "xlsx", "sheet": f"Sheet{idx}", "rows": rows[:60], "row_count": len(rows)})
        return tables


def _xlsx_shared_strings(workbook: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in workbook.namelist():
        return []
    root = ET.fromstring(workbook.read("xl/sharedStrings.xml"))
    strings = []
    for item in root:
        texts = [node.text or "" for node in item.iter() if node.tag.endswith("}t")]
        strings.append("".join(texts))
    return strings


def _csv_table(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        rows = list(csv.reader(handle))[:100]
    return {"kind": "csv", "sheet": path.stem, "rows": rows, "row_count": len(rows)}


def smart_summary(text: str, *, max_sentences: int = 6) -> str:
    clean = re.sub(r"\s+", " ", text).strip()
    if not clean:
        return "No readable text found."
    sentences = re.split(r"(?<=[.!?])\s+", clean)
    keywords = _tags_from_text(clean)
    selected = []
    for sentence in sentences:
        if len(sentence) < 25:
            continue
        score = sum(1 for tag in keywords if tag.lower() in sentence.lower())
        selected.append((score, sentence))
    selected.sort(key=lambda item: (item[0], len(item[1])), reverse=True)
    picked = [sentence for _score, sentence in selected[:max_sentences]] or [clean[:700]]
    return "\n".join(f"- {sentence[:420].strip()}" for sentence in picked)


def _iter_safe_documents(limit: int = 500) -> list[Path]:
    files: list[Path] = []
    for root in _safe_roots():
        if not root.exists():
            continue
        for current, dirs, names in os.walk(root):
            dirs[:] = [name for name in dirs if name not in SKIP_DIRS and not name.startswith(".")]
            for name in names:
                path = Path(current) / name
                if path.suffix.lower() in SUPPORTED_EXTENSIONS:
                    files.append(path)
                    if len(files) >= limit:
                        return files
    return files


def _safe_roots() -> list[Path]:
    home = Path.home()
    return [home / "Downloads", home / "Documents", home / "Desktop", Path("D:/Grandpa")]


def _tokens(text: str) -> set[str]:
    return {token for token in re.findall(r"[a-z0-9]+", text.lower()) if len(token) > 1}


def _tags_from_text(text: str) -> list[str]:
    stop = {"the", "and", "for", "with", "from", "this", "that", "your", "into", "file", "document"}
    counts: dict[str, int] = {}
    for token in re.findall(r"[a-zA-Z][a-zA-Z0-9_-]{2,}", text.lower()):
        if token in stop:
            continue
        counts[token] = counts.get(token, 0) + 1
    return [word for word, _count in sorted(counts.items(), key=lambda item: item[1], reverse=True)[:12]]


def _title_from_text(text: str) -> str:
    for line in text.splitlines():
        clean = line.strip().strip("#").strip()
        if 4 <= len(clean) <= 160:
            return clean
    return ""


def _query_wants_type(query: str, doc_type: str) -> bool:
    lower = query.lower()
    aliases = {
        "xlsx": ("excel", "spreadsheet", "sheet", "xlsx"),
        "csv": ("csv", "spreadsheet", "table"),
        "pdf": ("pdf",),
        "docx": ("word", "docx", "resume", "document"),
        "pptx": ("powerpoint", "presentation", "slides", "pptx"),
    }
    return any(word in lower for word in aliases.get(doc_type, (doc_type,)))


def _query_wants_last_month(query: str, modified: float) -> bool:
    if "last month" not in query.lower():
        return False
    now = datetime.now()
    start = (now.replace(day=1) - timedelta(days=1)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return start.timestamp() <= modified < end.timestamp()


def _organization_category(row: dict[str, Any]) -> str:
    tags = set(row.get("tags", []))
    name = Path(row["path"]).name.lower()
    if "invoice" in tags or "invoice" in name:
        return "Invoices"
    if row["type"] in {"xlsx", "csv"}:
        return "Spreadsheets"
    if row["type"] == "pptx":
        return "Presentations"
    if "resume" in tags or "resume" in name:
        return "Resumes"
    return "Documents"


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")[:80]


def _date(timestamp: float, *, compact: bool = False) -> str:
    return datetime.fromtimestamp(timestamp).strftime("%Y%m%d" if compact else "%b %d, %Y")


__all__ = [
    "DocumentIndex",
    "DocumentResult",
    "SUPPORTED_EXTENSIONS",
    "diagnostics",
    "document_metadata",
    "extract_document_text",
    "extract_tables",
    "index_documents",
    "organization_plan",
    "search_documents",
    "smart_summary",
    "suggest_renames",
    "summarize_document",
]
